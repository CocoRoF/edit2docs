"""M4 — XLSX grid engine + number-format subset."""

from __future__ import annotations

import datetime
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from edit2docs.documents.number_format import format_cell_value
from edit2docs.documents.xlsx_pages import xlsx_to_page_svgs


def _wb_bytes(build) -> bytes:
    wb = Workbook()
    build(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestNumberFormat:
    def test_general(self):
        assert format_cell_value(142.0, "General") == "142"
        assert format_cell_value(None, "General") == ""

    def test_thousands_and_decimals(self):
        assert format_cell_value(1420000, "#,##0") == "1,420,000"
        assert format_cell_value(3.14159, "0.00") == "3.14"

    def test_percent(self):
        assert format_cell_value(0.62, "0.0%") == "62.0%"
        assert format_cell_value(0.5, "0%") == "50%"

    def test_currency(self):
        assert format_cell_value(1420000, "₩#,##0") == "₩1,420,000"
        assert format_cell_value(-5, "$#,##0") == "-$5"

    def test_date(self):
        assert format_cell_value(datetime.date(2026, 7, 5), "yyyy-mm-dd") == "2026-07-05"
        assert (
            format_cell_value(datetime.datetime(2026, 7, 5, 9, 30), "yyyy-mm-dd h:mm")
            == "2026-07-05 09:30"
        )

    def test_unknown_format_never_raises(self):
        assert format_cell_value(5, "??weird??") == "5"


class TestXlsxPages:
    def test_grid_with_gutters_and_addresses(self):
        def build(wb):
            ws = wb.active
            ws.title = "매출"
            ws.append(["구분", "금액"])
            ws.append(["국내", 1420000])
            ws["B2"].number_format = "₩#,##0"
            ws["A1"].font = Font(bold=True)
            ws["A1"].fill = PatternFill("solid", fgColor="D9E2F3")

        pages = xlsx_to_page_svgs(_wb_bytes(build))
        assert len(pages) == 1
        svg = pages[0]
        assert 'data-e2d-sheet="매출"' in svg
        assert 'data-e2d-cell="B2"' in svg
        assert "₩1,420,000" in svg  # number format + auto-fit width
        assert 'fill="#D9E2F3"' in svg  # cell fill
        assert 'font-weight="bold"' in svg
        assert ">A</text>" in svg and ">1</text>" in svg  # gutters

    def test_merged_cells_render_once(self):
        def build(wb):
            ws = wb.active
            ws.merge_cells("A1:B1")
            ws["A1"] = "병합"
            ws.append(["x", "y"])

        svg = xlsx_to_page_svgs(_wb_bytes(build))[0]
        assert svg.count('data-e2d-cell="A1"') == 1
        assert 'data-e2d-cell="B1"' not in svg  # covered cell skipped

    def test_wide_sheet_splits_into_column_bands(self):
        def build(wb):
            ws = wb.active
            ws.append([f"col{i}" for i in range(40)])
            for c in range(1, 41):
                ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 20

        pages = xlsx_to_page_svgs(_wb_bytes(build))
        assert len(pages) >= 2

    def test_multi_sheet_renders_all(self):
        def build(wb):
            wb.active.append(["one"])
            wb.create_sheet("Second").append(["two"])

        pages = xlsx_to_page_svgs(_wb_bytes(build))
        joined = "".join(pages)
        assert 'data-e2d-sheet="Second"' in joined
        assert len(pages) == 2
