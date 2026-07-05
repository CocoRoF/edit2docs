"""XLSX → per-page SVG grid renderer (native-render plan M4).

Renders every sheet as a print-style grid: real column widths / row
heights, merged cells, solid fills, font bold/color, number-formatted
values (documents/number_format.py), and column-letter / row-number
gutters. Sheets larger than one page split into column bands × row
bands, each band becoming its own page (Excel's own print order:
down, then over).

Addressing: every page carries ``data-e2d-sheet`` and each cell group
``data-e2d-cell="B3"`` — identical to ``xlsx_to_html`` / ``set_doc_text``.
"""

from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from edit2docs.documents.number_format import format_cell_value

# Landscape A4-ish canvas in px (96/in): comfortable for wide sheets.
_PAGE_W = 1122.0
_PAGE_H = 794.0
_MARGIN = 28.0
_GUTTER_W = 34.0   # row-number column
_GUTTER_H = 18.0   # column-letter row
_TITLE_H = 22.0
_DEFAULT_COL_PX = 64.0
_DEFAULT_ROW_PX = 20.0
_CELL_PAD = 3.0
_FONT_PX = 11.0 * 96 / 72
_MAX_ROWS = 2000
_MAX_COLS = 128
_MAX_PAGES = 200

_FONT_STACK = "'Noto Sans CJK KR', 'Malgun Gothic', 'Noto Sans', sans-serif"


def _esc(s: str) -> str:
    return (
        s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _f(v: float) -> str:
    out = f"{v:.2f}".rstrip("0").rstrip(".")
    return out or "0"


def _argb_to_hex(argb: Optional[str]) -> Optional[str]:
    if not argb or not isinstance(argb, str):
        return None
    argb = argb.strip()
    if len(argb) == 8:
        argb = argb[2:]
    if len(argb) != 6:
        return None
    if argb.upper() in ("000000",):  # openpyxl noise default on empty styles
        return None
    return "#" + argb.upper()


def _cell_fill_hex(cell) -> Optional[str]:
    try:
        fill = cell.fill
        if fill is not None and fill.patternType == "solid":
            return _argb_to_hex(getattr(fill.fgColor, "rgb", None))
    except Exception:  # noqa: BLE001
        pass
    return None


def _cell_font(cell) -> tuple[bool, Optional[str], float]:
    bold, color, size = False, None, _FONT_PX
    try:
        font = cell.font
        bold = bool(font.b)
        color = _argb_to_hex(getattr(font.color, "rgb", None) if font.color else None)
        if font.sz:
            size = float(font.sz) * 96 / 72
    except Exception:  # noqa: BLE001
        pass
    return bold, color, size


def xlsx_to_page_svgs(content: bytes) -> list[str]:
    """Render every sheet of an XLSX to per-page SVG strings."""
    wb = load_workbook(io.BytesIO(content))
    try:
        wb_data = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:  # noqa: BLE001
        wb_data = None

    pages: list[str] = []
    for ws in wb.worksheets:
        ws_data = wb_data[ws.title] if wb_data is not None and ws.title in wb_data.sheetnames else None
        pages.extend(_render_sheet(ws, ws_data))
        if len(pages) >= _MAX_PAGES:
            break
    if not pages:
        pages.append(
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{_f(_PAGE_W)}" '
            f'height="{_f(_PAGE_H)}" viewBox="0 0 {_f(_PAGE_W)} {_f(_PAGE_H)}">'
            f'<rect width="100%" height="100%" fill="#ffffff"/></svg>'
        )
    return pages[:_MAX_PAGES]


def _render_sheet(ws, ws_data) -> list[str]:
    n_rows = min(ws.max_row or 1, _MAX_ROWS)
    n_cols = min(ws.max_column or 1, _MAX_COLS)

    col_px = []
    for c in range(1, n_cols + 1):
        dim = ws.column_dimensions.get(get_column_letter(c))
        width = getattr(dim, "width", None) if dim is not None else None
        if width:
            col_px.append(width * 7.0 + 5.0)
        else:
            # Auto-fit (Excel shows #### / spills; a preview should just
            # size the column): widest formatted value, clamped.
            widest = _DEFAULT_COL_PX
            for r in range(1, n_rows + 1):
                cell = ws.cell(row=r, column=c)
                value = cell.value
                if ws_data is not None and isinstance(value, str) and value.startswith("="):
                    cached = ws_data.cell(row=r, column=c).value
                    if cached is not None:
                        value = cached
                text = format_cell_value(value, cell.number_format)
                if not text:
                    continue
                est = sum(
                    _FONT_PX if ord(ch) > 0x2E80 else _FONT_PX * 0.58 for ch in text
                ) + _CELL_PAD * 2 + 4
                widest = max(widest, est)
            col_px.append(min(widest, 220.0))
    row_px = []
    for r in range(1, n_rows + 1):
        dim = ws.row_dimensions.get(r)
        height = getattr(dim, "height", None) if dim is not None else None
        row_px.append(height * 96 / 72 if height else _DEFAULT_ROW_PX)

    # merge map: anchor -> (rowspan, colspan); covered -> skip
    anchors: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        anchors[(rng.min_row, rng.min_col)] = (
            rng.max_row - rng.min_row + 1,
            rng.max_col - rng.min_col + 1,
        )
        for rr in range(rng.min_row, rng.max_row + 1):
            for cc in range(rng.min_col, rng.max_col + 1):
                if (rr, cc) != (rng.min_row, rng.min_col):
                    covered.add((rr, cc))

    # column bands × row bands that fit the content box
    content_w = _PAGE_W - _MARGIN * 2 - _GUTTER_W
    content_h = _PAGE_H - _MARGIN * 2 - _GUTTER_H - _TITLE_H
    col_bands = _bands(col_px, content_w)
    row_bands = _bands(row_px, content_h)

    out: list[str] = []
    for c0, c1 in col_bands:
        for r0, r1 in row_bands:
            out.append(
                _render_band(
                    ws, ws_data, col_px, row_px, anchors, covered,
                    r0, r1, c0, c1,
                )
            )
    return out


def _bands(sizes: list[float], limit: float) -> list[tuple[int, int]]:
    """Split 0-based index range into [start, end) bands fitting *limit*."""
    bands: list[tuple[int, int]] = []
    start = 0
    acc = 0.0
    for i, size in enumerate(sizes):
        if acc + size > limit and i > start:
            bands.append((start, i))
            start = i
            acc = 0.0
        acc += size
    bands.append((start, len(sizes)))
    return bands


def _render_band(
    ws, ws_data, col_px, row_px, anchors, covered, r0, r1, c0, c1,
) -> str:
    parts: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" version="1.1" '
        f'width="{_f(_PAGE_W)}" height="{_f(_PAGE_H)}" '
        f'viewBox="0 0 {_f(_PAGE_W)} {_f(_PAGE_H)}">'
        f'<rect width="100%" height="100%" fill="#ffffff"/>'
        f'<g data-e2d-sheet="{_esc(ws.title)}">'
    ]
    # sheet title
    parts.append(
        f'<text x="{_f(_MARGIN)}" y="{_f(_MARGIN + 12)}" font-size="13" '
        f'font-weight="bold" fill="#333333" font-family="{_FONT_STACK}">'
        f"{_esc(ws.title)}</text>"
    )
    ox = _MARGIN + _GUTTER_W
    oy = _MARGIN + _TITLE_H + _GUTTER_H

    # column-letter gutter
    x = ox
    for c in range(c0, c1):
        w = col_px[c]
        parts.append(
            f'<rect x="{_f(x)}" y="{_f(oy - _GUTTER_H)}" width="{_f(w)}" '
            f'height="{_f(_GUTTER_H)}" fill="#F2F2F2" stroke="#C9C9C9" stroke-width="0.6"/>'
            f'<text x="{_f(x + w / 2)}" y="{_f(oy - 5)}" text-anchor="middle" '
            f'font-size="10" fill="#666666" font-family="{_FONT_STACK}">'
            f"{get_column_letter(c + 1)}</text>"
        )
        x += w
    # row-number gutter
    y = oy
    for r in range(r0, r1):
        h = row_px[r]
        parts.append(
            f'<rect x="{_f(ox - _GUTTER_W)}" y="{_f(y)}" width="{_f(_GUTTER_W)}" '
            f'height="{_f(h)}" fill="#F2F2F2" stroke="#C9C9C9" stroke-width="0.6"/>'
            f'<text x="{_f(ox - _GUTTER_W / 2)}" y="{_f(y + h / 2 + 3.5)}" '
            f'text-anchor="middle" font-size="10" fill="#666666" '
            f'font-family="{_FONT_STACK}">{r + 1}</text>'
        )
        y += h

    # cells
    y = oy
    for r in range(r0, r1):
        x = ox
        for c in range(c0, c1):
            w, h = col_px[c], row_px[r]
            row_i, col_i = r + 1, c + 1
            if (row_i, col_i) in covered:
                x += w
                continue
            span = anchors.get((row_i, col_i))
            cw = sum(col_px[c:c + span[1]]) if span else w
            ch = sum(row_px[r:r + span[0]]) if span else h

            cell = ws.cell(row=row_i, column=col_i)
            value = cell.value
            if ws_data is not None and isinstance(value, str) and value.startswith("="):
                cached = ws_data.cell(row=row_i, column=col_i).value
                if cached is not None:
                    value = cached
            text = format_cell_value(value, cell.number_format)
            fill = _cell_fill_hex(cell)
            bold, color, size = _cell_font(cell)

            ref = f"{get_column_letter(col_i)}{row_i}"
            attrs = f' fill="{fill}"' if fill else ' fill="none"'
            cell_parts = [
                f'<g data-e2d-cell="{ref}">'
                f'<rect x="{_f(x)}" y="{_f(y)}" width="{_f(cw)}" height="{_f(ch)}"'
                f'{attrs} stroke="#C9C9C9" stroke-width="0.6"/>'
            ]
            if text:
                is_num = isinstance(value, (int, float)) and not isinstance(value, bool)
                max_chars = max(int((cw - _CELL_PAD * 2) / (size * 0.55)), 1)
                shown = text if len(text) <= max_chars else text[: max(max_chars - 1, 1)] + "…"
                tx = x + cw - _CELL_PAD if is_num else x + _CELL_PAD
                anchor = "end" if is_num else "start"
                weight = ' font-weight="bold"' if bold else ""
                fill_attr = color or "#222222"
                cell_parts.append(
                    f'<text x="{_f(tx)}" y="{_f(y + ch / 2 + size * 0.32)}" '
                    f'text-anchor="{anchor}" font-size="{_f(size)}" '
                    f'fill="{fill_attr}"{weight} font-family="{_FONT_STACK}" '
                    f'xml:space="preserve">{_esc(shown)}</text>'
                )
            cell_parts.append("</g>")
            parts.append("".join(cell_parts))
            x += w
        y += row_px[r]

    parts.append("</g></svg>")
    return "".join(parts)
