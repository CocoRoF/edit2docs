"""Deterministic XLSX building blocks (no LLM).

A *sheet spec* is the interchange format the generator LLM emits:

    sheets:
      - name: "매출 요약"
        headers: ["분기", "매출(억원)", "YoY"]
        rows:
          - ["1분기", 120, "+12%"]
          - ["2분기", 135, "+9%"]
        widths: [10, 14, 10]        # optional, characters
        number_formats: {"B": "#,##0"}   # optional, column letter -> format

``xlsx_from_spec`` renders it (styled header row, frozen panes, auto
widths); ``xlsx_outline`` / ``apply_xlsx_edits`` give agents the same
inspect-then-patch loop the DOCX/PPTX sides have.
"""

from __future__ import annotations

import io
import re
from collections.abc import Iterable
from dataclasses import dataclass
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

__all__ = [
    "xlsx_from_spec",
    "xlsx_outline",
    "xlsx_to_markdown",
    "xlsx_to_html",
    "xlsx_preview",
    "apply_xlsx_edits",
    "XlsxEdit",
    "XlsxEditResult",
]

_HEADER_FILL = PatternFill("solid", fgColor="1B64DA")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ---------------------------------------------------------------------------
# Spec -> XLSX
# ---------------------------------------------------------------------------


def xlsx_from_spec(spec: dict[str, Any]) -> bytes:
    """Render a sheet spec into a styled .xlsx package.

    Raises ``ValueError`` on a structurally invalid spec (bilingual
    message) — the generator retries on that signal.
    """
    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise ValueError(
            "sheet spec must contain a non-empty `sheets` list. "
            "sheet spec에는 비어있지 않은 `sheets` 목록이 필요합니다."
        )

    wb = Workbook()
    wb.remove(wb.active)
    seen_titles: set[str] = set()
    for index, sheet in enumerate(sheets, start=1):
        if not isinstance(sheet, dict):
            raise ValueError(f"sheets[{index - 1}] must be a mapping")
        title = str(sheet.get("name") or f"Sheet{index}")[:31]
        if title in seen_titles:
            raise ValueError(
                f"duplicate sheet name {title!r} (after 31-char truncation). "
                f"시트 이름 {title!r} 이 중복됩니다."
            )
        seen_titles.add(title)
        ws = wb.create_sheet(title=title)

        headers = [str(h) for h in (sheet.get("headers") or [])]
        rows = sheet.get("rows") or []
        if headers:
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _BORDER
            ws.freeze_panes = "A2"
        for row in rows:
            if not isinstance(row, (list, tuple)):
                raise ValueError(f"rows entries must be lists (sheet {title!r})")
            ws.append(list(row))

        # Body borders.
        first_body = 2 if headers else 1
        for row_cells in ws.iter_rows(min_row=first_body):
            for cell in row_cells:
                cell.border = _BORDER

        # Column widths: explicit, else content-driven (capped).
        widths = sheet.get("widths") or []
        column_count = max(len(headers), *(len(r) for r in rows), 1) if rows or headers else 1
        for col in range(1, column_count + 1):
            letter = get_column_letter(col)
            if col <= len(widths) and widths[col - 1]:
                try:
                    ws.column_dimensions[letter].width = float(widths[col - 1])
                except (TypeError, ValueError):
                    raise ValueError(
                        f"widths entries must be numbers (got {widths[col - 1]!r}). "
                        "widths 항목은 숫자여야 합니다."
                    ) from None
            else:
                longest = 0
                for row_cells in ws.iter_rows(min_col=col, max_col=col):
                    value = row_cells[0].value
                    if value is not None:
                        # CJK chars are ~2 cells wide.
                        text = str(value)
                        width = sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)
                        longest = max(longest, width)
                ws.column_dimensions[letter].width = min(max(longest + 2, 8), 48)

        from openpyxl.utils import column_index_from_string

        for letter, fmt in (sheet.get("number_formats") or {}).items():
            if not re.fullmatch(r"[A-Za-z]{1,3}", str(letter)):
                raise ValueError(
                    f"number_formats keys must be column letters (got {letter!r}). "
                    "number_formats 키는 열 문자(A, B, ...)여야 합니다."
                )
            col_idx = column_index_from_string(str(letter).upper())
            if col_idx > column_count:
                continue  # format for a column with no data — skip, don't create it
            start = 2 if headers else 1
            for (cell,) in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=start):
                cell.number_format = str(fmt)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XLSX -> outline / markdown
# ---------------------------------------------------------------------------


def _load(content: bytes):
    try:
        return load_workbook(io.BytesIO(content), data_only=False)
    except Exception as exc:
        raise ValueError(
            f"XLSX could not be opened: {exc}. XLSX 파일을 열 수 없습니다."
        ) from exc


def xlsx_outline(content: bytes, *, sample_rows: int = 8) -> dict:
    """Workbook structure: sheets, dimensions, headers, sample rows.

    Additive: a workbook-level ``"charts"`` list (``kind`` / ``title`` /
    ``series_names`` per chart, read via contextifier's ChartModel) so
    the edit planner is not blind to charts it must not clobber.
    """
    wb = _load(content)
    sheets = []
    for ws in wb.worksheets:
        rows = ws.max_row if ws.max_row else 0
        cols = ws.max_column if ws.max_column else 0
        sample = []
        for row_cells in ws.iter_rows(min_row=1, max_row=min(rows, sample_rows)):
            sample.append([cell.value for cell in row_cells])
        sheets.append(
            {
                "name": ws.title,
                "rows": rows,
                "columns": cols,
                "sample": sample,
            }
        )
    return {"sheets": sheets, "charts": _chart_outline(content)}


def _chart_outline(content: bytes) -> list[dict]:
    """Read-only chart summaries via contextifier (best-effort: outline
    must never fail because a chart part is exotic)."""
    try:
        from contextifier import open_raw

        raw = open_raw(content, extension="xlsx")
        return [
            {
                "kind": chart.kind,
                "title": chart.title,
                "series_names": [s.name for s in chart.series],
            }
            for chart in raw.charts
        ]
    except Exception:
        return []


def xlsx_to_markdown(content: bytes, *, max_rows: int = 40) -> str:
    """Every sheet as a markdown table (row-capped for prompt budgets)."""
    wb = _load(content)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"## {ws.title}")
        rows = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, max_rows)))
        if not rows:
            parts.append("(empty)")
            continue
        table_lines = []
        for i, row_cells in enumerate(rows):
            cells = ["" if c.value is None else str(c.value) for c in row_cells]
            table_lines.append("| " + " | ".join(cells) + " |")
            if i == 0:
                table_lines.append("|" + "---|" * len(cells))
        if (ws.max_row or 0) > max_rows:
            table_lines.append(f"… ({ws.max_row - max_rows} more rows)")
        parts.append("\n".join(table_lines))
    return "\n\n".join(parts)


# ---------------------------------------------------------------------------
# Targeted edits
# ---------------------------------------------------------------------------

_CELL_REF = re.compile(r"^[A-Za-z]{1,3}[1-9]\d*$")


@dataclass
class XlsxEdit:
    """One operation. ``action``:

    * ``set_cell`` — write ``value`` into ``cell`` (e.g. "B3") of ``sheet``.
      ``old_value`` (stringified compare) guards staleness.
    * ``append_rows`` — append ``rows`` (list of lists) to ``sheet``.
    * ``add_sheet`` — create sheet ``sheet`` with optional ``headers``/``rows``.
    """

    action: str
    sheet: str = ""
    cell: str | None = None
    value: Any = None
    old_value: Any = None
    rows: list | None = None
    headers: list | None = None


@dataclass
class XlsxEditResult:
    action: str
    status: str  # applied | stale | not_found | invalid
    message: str = ""


def apply_xlsx_edits(content: bytes, edits: Iterable[XlsxEdit]) -> tuple[bytes, list[XlsxEditResult]]:
    """Apply edits losslessly via contextifier's raw layer.

    Only the worksheet parts an edit actually touches (plus
    ``xl/workbook.xml`` when a formula cache goes stale) are rewritten —
    charts, chart styles, sparklines, pivot tables, customXml and cached
    formula values all survive byte-identical (the old openpyxl
    load→save round-trip destroyed every one of those on EVERY edit).
    """
    from contextifier import open_raw

    try:
        raw = open_raw(content, extension="xlsx")
    except Exception as exc:
        raise ValueError(
            f"XLSX could not be opened: {exc}. XLSX 파일을 열 수 없습니다."
        ) from exc
    results: list[XlsxEditResult] = []
    for edit in edits:
        results.append(_apply_one(raw, edit))
    return raw.to_bytes(), results


def _apply_one(raw, edit: XlsxEdit) -> XlsxEditResult:
    if edit.action == "add_sheet":
        if not edit.sheet:
            return XlsxEditResult(edit.action, "invalid", "add_sheet needs a sheet name")
        if edit.sheet in raw.sheet_names:
            return XlsxEditResult(edit.action, "invalid", f"sheet {edit.sheet!r} already exists")
        title = str(edit.sheet)[:31]
        rows: list[list] = []
        if edit.headers:
            rows.append(list(edit.headers))
        for row in edit.rows or []:
            if not isinstance(row, (list, tuple)):
                return XlsxEditResult(edit.action, "invalid", "rows entries must be lists")
            rows.append(list(row))
        try:
            _add_raw_sheet(raw, title)
            if rows:
                raw.sheets[title].append_rows(rows)
        except (TypeError, ValueError) as exc:
            return XlsxEditResult(edit.action, "invalid", str(exc))
        return XlsxEditResult(edit.action, "applied")

    if edit.sheet not in raw.sheet_names:
        return XlsxEditResult(
            edit.action, "not_found",
            f"sheet {edit.sheet!r} not in {raw.sheet_names}",
        )
    sheet = raw.sheets[edit.sheet]

    if edit.action == "set_cell":
        if not edit.cell or not _CELL_REF.match(edit.cell):
            return XlsxEditResult(edit.action, "invalid", f"bad cell ref {edit.cell!r}")
        from openpyxl.utils import column_index_from_string

        col_letters = re.match(r"^[A-Za-z]+", edit.cell).group(0)
        if column_index_from_string(col_letters.upper()) > 16384:
            return XlsxEditResult(
                edit.action, "invalid", f"column {col_letters!r} beyond Excel's XFD limit"
            )
        ref = edit.cell.upper()
        if edit.old_value is not None:
            value = sheet.get_cell(ref)
            current = "" if value is None else str(value)
            if current.strip() != str(edit.old_value).strip():
                return XlsxEditResult(edit.action, "stale", "cell value changed; refresh")
        try:
            sheet.set_cell(ref, edit.value)
        except (TypeError, ValueError) as exc:
            return XlsxEditResult(edit.action, "invalid", str(exc))
        return XlsxEditResult(edit.action, "applied")

    if edit.action == "append_rows":
        if not edit.rows:
            return XlsxEditResult(edit.action, "invalid", "append_rows needs rows")
        for row in edit.rows:
            if not isinstance(row, (list, tuple)):
                return XlsxEditResult(edit.action, "invalid", "rows entries must be lists")
        try:
            sheet.append_rows([list(row) for row in edit.rows])
        except (TypeError, ValueError) as exc:
            return XlsxEditResult(edit.action, "invalid", str(exc))
        return XlsxEditResult(edit.action, "applied")

    return XlsxEditResult(edit.action, "invalid", f"unknown action {edit.action!r}")


_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_WORKSHEET_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
)
# A schema-minimal worksheet (CT_Worksheet only requires <sheetData>) —
# the same skeleton an openpyxl-written sheet reduces to.
_MINIMAL_WORKSHEET_XML = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
    b"<sheetData/></worksheet>"
)


def _add_raw_sheet(raw, title: str) -> None:
    """Create an empty worksheet inside the raw package.

    The raw layer has no ``add_sheet``, so wire the four OPC touchpoints
    here: worksheet part, ``[Content_Types].xml`` override, workbook
    rels entry, and the ``<sheet>`` element in ``xl/workbook.xml``.
    Everything else in the package stays byte-identical.
    """
    from contextifier.raw import qn

    package = raw.package
    n = 1
    while package.has_part(f"xl/worksheets/sheet{n}.xml"):
        n += 1
    part_name = f"xl/worksheets/sheet{n}.xml"
    package.add_part(part_name, _MINIMAL_WORKSHEET_XML)
    package.set_content_type_override(part_name, _WORKSHEET_CONTENT_TYPE)

    workbook = raw.workbook  # XmlPart facade over xl/workbook.xml
    rels = package.rels_for(workbook.name)
    if rels is None:
        raise ValueError("workbook has no relationships part")
    rel_id = rels.next_id()
    rels.add(rel_id, _WORKSHEET_REL_TYPE, f"worksheets/sheet{n}.xml")

    sheets_el = workbook.find("s:sheets")
    if sheets_el is None:
        raise ValueError("workbook.xml has no <sheets> element")
    sheet_ids = [
        int(s.get("sheetId", "0"))
        for s in sheets_el.findall(qn("s:sheet"))
        if (s.get("sheetId") or "0").isdigit()
    ]
    sheet_el = sheets_el.makeelement(qn("s:sheet"), {})
    sheet_el.set("name", title)
    sheet_el.set("sheetId", str(max(sheet_ids, default=0) + 1))
    sheet_el.set(qn("r:id"), rel_id)
    sheets_el.append(sheet_el)
    workbook.mark_dirty()


def xlsx_preview(content: bytes, *, max_rows: int = 200) -> tuple[str, list[dict]]:
    """Addressable display HTML + warnings for the studio preview.

    Structure per sheet: ``<section class="e2d-sheet" data-e2d-sheet="이름">``
    with a spreadsheet-style grid — column-letter header row, row-number
    column, and every data cell stamped ``data-e2d-cell="B3"`` (the exact
    address ``set_cell`` edits take, mirroring the DOCX preview's
    ``data-e2d-para`` convention). Merged ranges render as real
    colspan/rowspan. Formula cells display their cached result when the
    file carries one (the formula moves to the ``title`` tooltip) instead
    of the raw ``=SUM(...)`` text the old preview showed.
    """
    from html import escape

    wb = _load(content)
    warnings: list[dict] = []
    try:  # cached formula results live in a second, values-only read
        wb_values = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        wb_values = None

    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f'<section class="e2d-sheet" data-e2d-sheet="{escape(ws.title, quote=True)}">')
        parts.append(f"<h2>{escape(ws.title)}</h2>")
        total = ws.max_row or 0
        cols = ws.max_column or 0
        rows = list(ws.iter_rows(min_row=1, max_row=min(total, max_rows)))
        if not rows:
            parts.append("<p>(empty)</p></section>")
            continue
        ws_values = None
        if wb_values is not None and ws.title in wb_values.sheetnames:
            ws_values = wb_values[ws.title]

        # Merged ranges: top-left renders with span; covered cells skip.
        spans: dict[tuple[int, int], tuple[int, int]] = {}
        covered: set[tuple[int, int]] = set()
        for rng in ws.merged_cells.ranges:
            spans[(rng.min_row, rng.min_col)] = (
                rng.max_row - rng.min_row + 1,
                rng.max_col - rng.min_col + 1,
            )
            for rr in range(rng.min_row, rng.max_row + 1):
                for cc in range(rng.min_col, rng.max_col + 1):
                    if (rr, cc) != (rng.min_row, rng.min_col):
                        covered.add((rr, cc))

        # Header heuristic: an explicit freeze line below row 1, or an
        # all-text first row over data rows.
        first_row = [c.value for c in rows[0]]
        frozen_header = bool(ws.freeze_panes and str(ws.freeze_panes)[1:].isdigit()
                             and int(str(ws.freeze_panes)[1:] or 1) > 1)
        texty_header = (
            len(rows) > 1
            and any(v is not None for v in first_row)
            and all(isinstance(v, str) for v in first_row if v is not None)
        )
        header_rows = 1 if (frozen_header or texty_header) else 0

        body = ['<table class="e2d-grid">']
        # Column-letter header row (A, B, C…) with a corner cell.
        letters = "".join(
            f'<th class="e2d-colhead">{get_column_letter(ci)}</th>'
            for ci in range(1, cols + 1)
        )
        body.append(f'<tr><th class="e2d-corner"></th>{letters}</tr>')
        for row_cells in rows:
            row_number = row_cells[0].row if row_cells else 1
            cells_html = [f'<th class="e2d-rowhead">{row_number}</th>']
            for cell in row_cells:
                key = (cell.row, cell.column)
                if key in covered:
                    continue
                span_attr = ""
                if key in spans:
                    rowspan, colspan = spans[key]
                    if rowspan > 1:
                        span_attr += f' rowspan="{rowspan}"'
                    if colspan > 1:
                        span_attr += f' colspan="{colspan}"'
                value = cell.value
                title_attr = ""
                if isinstance(value, str) and value.startswith("="):
                    title_attr = f' title="{escape(value, quote=True)}"'
                    if ws_values is not None:
                        cached = ws_values.cell(row=cell.row, column=cell.column).value
                        if cached is not None:
                            value = cached
                css_class = ""
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    css_class = ' class="e2d-num"'
                # M4: honour the cell's Excel number format (currency,
                # percent, thousands, dates) instead of raw repr.
                from edit2docs.documents.number_format import format_cell_value

                text = escape(format_cell_value(value, cell.number_format))
                tag = "th" if cell.row <= header_rows else "td"
                cells_html.append(
                    f'<{tag} data-e2d-cell="{cell.coordinate}"'
                    f"{span_attr}{title_attr}{css_class}>{text}</{tag}>"
                )
            body.append(f"<tr>{''.join(cells_html)}</tr>")
        body.append("</table>")
        if total > max_rows:
            body.append(f"<p>… ({total - max_rows} more rows)</p>")
            warnings.append(
                {
                    "code": "preview_rows_truncated",
                    "message": (
                        f"Sheet {ws.title!r}: {total} rows; preview shows the "
                        f"first {max_rows}. 미리보기는 앞 {max_rows}행만 표시합니다."
                    ),
                }
            )
        parts.append("".join(body))
        parts.append("</section>")
    return "\n".join(parts), warnings


def xlsx_to_html(content: bytes, *, max_rows: int = 200) -> str:
    """Every sheet as an HTML grid (row-capped) — see :func:`xlsx_preview`."""
    html, _warnings = xlsx_preview(content, max_rows=max_rows)
    return html
